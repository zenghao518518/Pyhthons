import requests
from openpyxl import workbook
# 1.打开Excel和Sheet
wb = workbook.Workbook()
sheet = wb.worksheets[0]
res = requests.get("http://www.10086.cn/support/selfservice/help/sh/5010801_4073_8801.json")
data_dict = res.json()
data_list = data_dict["cData"]['list']
row_index = 1
for item in data_list:
    question = item['question']
    up_time = item['up_time']
    id = item['_orderId']
    sheet.cell(row_index, 1).value = id
    sheet.cell(row_index, 2).value = question
    sheet.cell(row_index, 3).value = up_time
    row_index = row_index + 1
wb.save("new.xlsx")
