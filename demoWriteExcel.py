# pip install openpyxl
from openpyxl import workbook

# 创建excel且默认会创建一个sheet (名称为Sheet)
wb = workbook.Workbook()
sheet = wb.worksheets[0]  # 或sheet = wb[" Sheet"]
# 找到单元格，并修改单元格的内容
cell = sheet.cell(1, 1)
cell.value = "新的开始"
sheet.cell(1, 2).value = "start"
# 将excel文件保存到p2.xlsx文件中
wb.save(" p2.xlsx")
